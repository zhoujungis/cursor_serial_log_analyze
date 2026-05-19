# -*- coding: utf-8 -*-
"""
桌面 GUI（PyQt6）：打开串口/设备 .log/.txt，规则扫描 + Cloud 清洗节选 + LLM 总结。
清洗文案、节选/材料体量：菜单「设置」；串口匹配规则与表格导入：菜单「规则」。依赖：requests、python-dotenv、pypdf（仅 .pdf）、openpyxl（.xlsx/.xlsm）、csv（.csv 规则导入）。

运行：python desktop_serial_log_analyzer.py
"""
from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import queue
import re
import sys
import threading
import traceback
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

import requests
from dotenv import load_dotenv

from utils.serial_alert_rules import (
    RAW_ALERT_RULE_DEFINITIONS,
    load_user_rules_raw,
    save_user_rules_raw,
)

_BUILTIN_RULE_CATEGORIES: frozenset[str] = frozenset(t[1] for t in RAW_ALERT_RULE_DEFINITIONS)


def _app_root() -> Path:
    """源码：项目目录；PyInstaller exe：exe 所在目录（.env / 用户规则与输出持久化）。"""
    if getattr(sys, "frozen", False) and getattr(sys, "executable", None):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent


_ROOT = _app_root()
load_dotenv(_ROOT / "config/.env")
load_dotenv(Path.cwd() / "config/.env")

# ----- 内联：日志解析 + LLM Cloud（原独立 summarize_*.py）-----

_INTERESTING = re.compile(
    r"(ERROR|ERR\b|FATAL|CRITICAL|\bWARN(?:ING)?\b|"
    r"异常|错误|失败|告警|Traceback|\bException\b)",
    re.I,
)
_TS_HEAD = re.compile(
    r"^(\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}(?:\.\d+)?|\d{4}/\d{2}/\d{2}\s+\d{2}:\d{2}:\d{2})\s*"
)


@dataclass
class Incident:
    seq: int
    ts: str
    source_file: str
    tags: list[str]
    log_line: str
    analysis: str = ""


def _read_file_text(path: Path) -> str:
    suf = path.suffix.lower()
    if suf == ".pdf":
        from pypdf import PdfReader

        reader = PdfReader(str(path))
        chunks: list[str] = []
        for page in reader.pages:
            chunks.append(page.extract_text() or "")
        return "\n".join(chunks)
    raw = path.read_bytes()
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _line_ts(line: str) -> str:
    m = _TS_HEAD.match(line.strip())
    return m.group(1).strip() if m else ""


def _line_tags(line: str) -> list[str]:
    tags: list[str] = []
    low = line.lower()
    if "traceback" in low:
        tags.append("TRACEBACK")
    if re.search(r"\bexception\b", low):
        tags.append("EXCEPTION")
    for kw in ("fatal", "critical", "error", "err", "warning", "warn"):
        if re.search(rf"\b{re.escape(kw)}\b", low):
            tags.append(kw.upper())
    for zh in ("异常", "错误", "失败", "告警"):
        if zh in line:
            tags.append(zh)
    return tags


def bug_type_cn(tags: list[str]) -> str:
    blob = " ".join(tags).lower()
    if "timeout" in blob or "超时" in blob:
        return "超时/无响应"
    if "memory" in blob or "oom" in blob or "内存" in blob:
        return "内存/资源"
    if "network" in blob or "socket" in blob or "连接" in blob:
        return "网络/连接"
    if "traceback" in blob or "exception" in blob:
        return "异常/栈"
    if "error" in blob or "err" in blob or "错误" in blob or "异常" in blob:
        return "运行错误"
    if "warn" in blob or "warning" in blob or "告警" in blob:
        return "告警"
    return "其他"


def severity_from_tags(tags: list[str], log_line: str) -> str:
    blob = (" ".join(tags) + " " + log_line).upper()
    if any(x in blob for x in ("FATAL", "CRITICAL", "严重", "崩溃")):
        return "高"
    if any(
        x in blob
        for x in ("ERROR", " ERR", "EXCEPTION", "TRACEBACK", "错误", "异常", "失败")
    ):
        return "中"
    if any(x in blob for x in ("WARN", "WARNING", "告警")):
        return "低"
    return "提示"


def short_title(log_line: str, max_len: int = 72) -> str:
    s = (log_line or "").strip().replace("\n", " ")
    if len(s) <= max_len:
        return s
    return s[: max_len - 1] + "…"


def parse_report_text(text: str, source_file: str) -> list[Incident]:
    out: list[Incident] = []
    seq = 0
    for line in text.splitlines():
        line_r = line.rstrip("\n\r")
        if not line_r.strip():
            continue
        if not _INTERESTING.search(line_r):
            continue
        seq += 1
        tags = _line_tags(line_r)
        if not tags:
            tags = ["SIGNAL"]
        out.append(
            Incident(
                seq=seq,
                ts=_line_ts(line_r),
                source_file=source_file,
                tags=tags,
                log_line=line_r.strip(),
                analysis="",
            )
        )
    return out


DEEPSEEK_API_BASE = "https://api.deepseek.com"


def _T(*parts: str) -> str:
    return "".join(parts)


def _repair_mojibake_text(text: str) -> str:
    if not isinstance(text, str) or not text:
        return text
    mojibake_hits = ("Ã", "Â", "æ", "å", "ä")
    if not any(ch in text for ch in mojibake_hits):
        return text
    try:
        fixed = text.encode("latin1", errors="strict").decode("utf-8", errors="strict")
        if fixed.count(" ") <= text.count(" "):
            return fixed
    except Exception:
        return text
    return text


def _deepseek_submit(prompt: str) -> str:
    """调用 DeepSeek API，返回完整响应文本。"""
    api_key = os.environ.get("DEEPSEEK_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("请在 .env 中设置 DEEPSEEK_API_KEY")

    model = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat").strip() or "deepseek-chat"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    body = {
        "model": model,
        "messages": [{"role": "user", "content": prompt}],
        "stream": False,
        "max_tokens": 8192,
    }
    resp = requests.post(
        f"{DEEPSEEK_API_BASE}/v1/chat/completions",
        headers=headers,
        json=body,
        timeout=600,
    )
    resp.raise_for_status()
    result = resp.json()
    return result["choices"][0]["message"]["content"]


def _xlsx_workbook_to_plain(path: Path, max_chars: int = 36_000) -> str:
    """将工作簿各表导出为可读纯文本（制表符分列），供 LLM 提取串口匹配规则。"""
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError("缺少 openpyxl，请执行：pip install openpyxl") from e
    wb = load_workbook(str(path), read_only=False, data_only=True)
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"\n## 工作表：{sheet_name}\n")
            for row in ws.iter_rows(values_only=True):
                cells: list[str] = []
                for c in row or ():
                    if c is None:
                        cells.append("")
                    else:
                        s = str(c).strip().replace("\r\n", "\n").replace("\r", "\n")
                        if "\n" in s:
                            s = re.sub(r"\s+", " ", s)
                        cells.append(s)
                if any(x for x in cells):
                    line = "\t".join(cells).strip()
                    if line:
                        parts.append(line + "\n")
    finally:
        wb.close()
    blob = "".join(parts).strip()
    if not blob:
        raise RuntimeError("Excel 中无有效文本单元格。")
    if len(blob) > max_chars:
        blob = (
            blob[:max_chars].rstrip()
            + "\n\n（以上内容因长度限制已截断，请仅基于已给片段归纳规则。）\n"
        )
    return blob


def _csv_file_to_plain(path: Path, max_chars: int = 36_000) -> str:
    """将 CSV 解码为与 xlsx 导出风格一致的纯文本（制表符分列），供 LLM 提取规则。"""
    raw = path.read_bytes()
    text: str | None = None
    for enc in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            text = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        text = raw.decode("utf-8", errors="replace")

    sample = text[:8192]
    dialect: type[csv.Dialect] | csv.Dialect = csv.excel
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        first = sample.splitlines()[0] if sample else ""
        if first.count("\t") >= first.count(",") and "\t" in first:
            dialect = csv.excel_tab

    parts: list[str] = [f"\n## 文件：{path.name}（CSV）\n"]
    total = len(parts[0])
    f = io.StringIO(text)
    reader = csv.reader(f, dialect=dialect)
    for row in reader:
        cells: list[str] = []
        for c in row:
            s = (c or "").strip().replace("\r\n", "\n").replace("\r", "\n")
            if "\n" in s:
                s = re.sub(r"\s+", " ", s)
            cells.append(s)
        if not any(cells):
            continue
        line = "\t".join(cells).strip()
        if not line:
            continue
        chunk = line + "\n"
        if total + len(chunk) > max_chars:
            parts.append("\n（以上内容因长度限制已截断，请仅基于已给片段归纳规则。）\n")
            break
        parts.append(chunk)
        total += len(chunk)

    blob = "".join(parts).strip()
    body = blob
    marker = f"## 文件：{path.name}（CSV）"
    if marker in body:
        body = body.split(marker, 1)[-1].strip()
    if not body:
        raise RuntimeError("CSV 中无有效数据行。")
    return blob


def _tabular_rules_source_to_plain(path: Path, max_chars: int = 36_000) -> str:
    suf = path.suffix.lower()
    if suf == ".csv":
        return _csv_file_to_plain(path, max_chars)
    if suf in (".xlsx", ".xlsm"):
        return _xlsx_workbook_to_plain(path, max_chars)
    raise RuntimeError("仅支持 .xlsx、.xlsm 或 .csv 文件。")


_RULE_IMPORT_INSTRUCTIONS = (
    "你是嵌入式串口日志匹配规则设计助手。下面文本来自「压测/稳定性测试用例」等表格："
    "Excel 多工作表导出，或 CSV 按行分列；其中常含「主要日志」「异常日志」「关键日志」「关注的日志」等段落，"
    "以及用 | 分隔的关键字列表。\n\n"
    "任务：从中提炼若干条用于在设备串口 log 单行文本上做正则匹配的告警规则。\n"
    "硬性要求：\n"
    "1) 只输出一个 JSON 数组本体，不要用 Markdown 代码围栏，不要附加解释性文字。\n"
    "2) 数组元素为对象，字段：priority（整数，默认 5，数字越小越优先）、"
    "category（英文小写与下划线组成的唯一键，如 ble_bind_fail）、"
    "label（中文简短显示名）、pattern（Python re 可用的正则字符串；默认忽略大小写，"
    "若未写 (?i) 前缀则仍按忽略大小写编译）。\n"
    "3) 每条规则聚焦一类现象；可把文档里用 | 列举的同组关键词合并为一条 alternation 正则。\n"
    "4) 总条数不超过 48；不要编造文档未提及的现象。\n"
    "5) pattern 必须合法，避免无转义的裸反斜杠错误；优先用非捕获组 (?:...) 控制分支。\n\n"
    "示例（仅说明格式，勿照抄）：\n"
    '[{"priority":5,"category":"mqtt_publish_err","label":"MQTT发布失败",'
    '"pattern":"(?i)(mqtt.+(fail|err)|publish.+fail)"}]\n\n'
)


def _rule_import_prompt(table_plain: str) -> str:
    return _RULE_IMPORT_INSTRUCTIONS + "--- 表格导出 ---\n\n" + table_plain


def _parse_json_array_from_llm(text: str) -> list:
    t = (text or "").strip()
    if not t:
        raise ValueError("模型返回为空。")
    m = re.search(r"```(?:json)?\s*([\s\S]*?)\s*```", t, re.I)
    if m:
        t = m.group(1).strip()
    start = t.find("[")
    if start < 0:
        raise ValueError("未在模型输出中找到 JSON 数组（以 [ 开头）。")
    dec = json.JSONDecoder()
    try:
        data, _end = dec.raw_decode(t, start)
    except json.JSONDecodeError as e:
        raise ValueError(f"JSON 解析失败：{e}") from e
    if not isinstance(data, list):
        raise ValueError("顶层 JSON 须为数组。")
    return data


def _slug_rule_category(raw: str, fallback: str) -> str:
    s = (raw or "").strip().lower()
    s = re.sub(r"[^a-z0-9_]+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if not s:
        fb = re.sub(r"[^a-z0-9_]+", "_", (fallback or "").lower())
        fb = re.sub(r"_+", "_", fb).strip("_")[:32]
        if fb:
            s = fb
        else:
            h = hashlib.md5((fallback or "rule").encode("utf-8")).hexdigest()[:10]
            s = "case_" + h
    return s[:64]


def _coerce_imported_rule_dict(obj: object, seq: int) -> dict:
    if not isinstance(obj, dict):
        raise ValueError(f"第 {seq} 条规则不是 JSON 对象。")
    cat_raw = str(obj.get("category", "")).strip()
    pat = str(obj.get("pattern", "")).strip()
    lbl = str(obj.get("label", "")).strip() or cat_raw or f"规则{seq}"
    pri_raw = obj.get("priority", 5)
    try:
        pri = int(pri_raw)
    except (TypeError, ValueError):
        pri = 5
    if not pat:
        raise ValueError(f"第 {seq} 条缺少 pattern。")
    try:
        re.compile(pat, re.I)
    except re.error as e:
        raise ValueError(f"第 {seq} 条正则无效：{e}") from e
    cat = _slug_rule_category(cat_raw, lbl)
    return {"priority": pri, "category": cat, "label": lbl, "pattern": pat}


def _rules_from_llm_response(text: str) -> list[dict]:
    arr = _parse_json_array_from_llm(text)
    out: list[dict] = []
    for i, it in enumerate(arr, start=1):
        out.append(_coerce_imported_rule_dict(it, i))
    return out


def _incident_block(inc: Incident) -> str:
    sev = severity_from_tags(inc.tags, inc.log_line)
    btype = bug_type_cn(inc.tags)
    title = short_title(inc.log_line)
    ana = inc.analysis.strip() if inc.analysis else ""
    lines = [
        f"- [{inc.seq}] ts={inc.ts} file={inc.source_file}",
        f"  type={btype} sev_hint={sev} title_hint={title}",
        f"  log: {inc.log_line}",
    ]
    if ana:
        lines.append(f"  analysis: {ana}")
    return "\n".join(lines)


def build_material_for_prompt(
    incidents: list[Incident],
    max_chars: int,
) -> tuple[str, bool]:
    header = "## Parsed incidents (structured)\n\n"
    total = header
    truncated = False
    for inc in incidents:
        block = _incident_block(inc) + "\n\n"
        if len(total) + len(block) > max_chars:
            truncated = True
            break
        total += block
    return total.strip(), truncated


def build_summarize_prompt_serial(
    material: str,
    log_files: list[str],
    truncated: bool,
    structured: bool,
) -> str:
    note_trunc = ""
    if truncated:
        note_trunc = (
            "\n"
            "注意：材料已截断，请在文首说明「材料已截断」并基于已给内容总结。\n"
        )
    mat_desc = (
        "结构化告警/分析条目"
        if structured
        else "原文节选（未能解析行号格式）"
    )
    return _T(
        "你是嵌入式/设备日志分析专家。以下为从串口原始日志文件中提取的",
        mat_desc,
        "。\n\n",
        "涉及的文件：\n",
        "\n".join(f"- {x}" for x in log_files),
        note_trunc,
        "\n\n",
        "请仅用简体中文输出纯文本。按下列结构撰写便于拷贝到 Bug 单/邮件：\n",
        "\n1) 概要一段\n",
        "2) 按严重程度归纳（合并同类）\n",
        "3) Bug 清单每条款必含字段：",
        "【Bug编号】 ",
        "【标题】 ",
        "【严重级别】 ",
        "【类型】 ",
        "【时间】 ",
        "【证据/原文】 ",
        "【分析结论】\n",
        "4) 后续排查顺序（序号列表）\n\n",
        "--- material ---\n",
        material,
    )


_USER_RULES_PATH = _ROOT / "config/serial_rules_user.json"

APP_VERSION = "V 1.0"
APP_AUTHOR = "zhoujun@glazero.com"
DISCLAIMER_TEXT = (
    "本次分析结果只对本次导入的串口 log 有效；\n"
    "具体结果需由测试人员复现、验证后再作结论。"
)


def apply_deepseek_env_to_dotenv(updates: dict[str, str]) -> None:
    """将 DeepSeek 相关变量写入项目根目录 .env（同名键覆盖；值为空则删除该键）。"""
    env_path = _ROOT / "config/.env"
    keys = set(updates.keys())
    kept: list[str] = []
    if env_path.exists():
        for raw in env_path.read_text(encoding="utf-8").splitlines():
            s = raw.strip()
            if not s or s.startswith("#"):
                kept.append(raw)
                continue
            if "=" in s:
                k = s.split("=", 1)[0].strip()
                if k in keys:
                    continue
            kept.append(raw)
    for k, v in updates.items():
        v = (v or "").strip()
        if v:
            kept.append(f"{k}={v}")
    env_path.parent.mkdir(parents=True, exist_ok=True)
    env_path.write_text("\n".join(kept).rstrip() + "\n", encoding="utf-8")
    for k, v in updates.items():
        vs = (v or "").strip()
        if vs:
            os.environ[k] = vs
        else:
            os.environ.pop(k, None)
    try:
        load_dotenv(env_path, override=True)
    except TypeError:
        load_dotenv(env_path)
    try:
        load_dotenv(Path.cwd() / "config/.env", override=True)
    except TypeError:
        load_dotenv(Path.cwd() / "config/.env")


def _default_cleaning_prompt() -> str:
    return (
        "你是嵌入式串口日志清洗助手。下面是一段原始日志节选（可能含噪声、重复前缀、ANSI 颜色码已尽量去除）。\n"
        "请用简体中文输出：\n"
        "1) 你认为的噪声类型简述；\n"
        "2) 对节选做「规范化重排」（合并明显重复心跳、保留关键时间顺序），输出重排后的正文，不要编造不存在的日志。\n"
        "仅基于给定节选作答。\n\n"
        "--- excerpt ---\n"
    )


@dataclass
class WorkerConfig:
    file_path: Path
    max_chars: int
    user_clean_prompt: str
    user_analysis_notes: str


def _run_analyze(cfg: WorkerConfig, q: queue.Queue) -> None:
    from utils.serial_alert_rules import build_compiled_rules, match_log_alerts_for_rules

    def _raw_material(text: str, budget: int, name: str) -> tuple[str, bool]:
        truncated = False
        head = "## raw excerpt: " + name + "\n\n" + text.strip()
        if len(head) > budget:
            material = head[:budget].strip()
            truncated = True
        else:
            material = head.strip()
        if not material:
            raise RuntimeError("无有效文本")
        return material, truncated

    try:
        # ══════════════════════════════════════════════════════════
        # 阶段 1：读取 + 规范化 + Cloud 清洗
        # ══════════════════════════════════════════════════════════
        q.put(("progress", 5, "读取文件…"))
        raw = _read_file_text(cfg.file_path)
        if not raw.strip():
            q.put(("err", "文件为空或无法解码。"))
            return

        q.put(("progress", 10, "规范化换行…"))
        cleaned = raw.replace("\r\n", "\n").replace("\r", "\n")

        q.put(("progress", 15, "调用 Cloud 清洗节选（前置）…"))
        try:
            clean_limit = int(os.environ.get("CLEAN_MAX_CHARS", "24000"))
        except ValueError:
            clean_limit = 24000
        clean_limit = max(1000, min(clean_limit, 80_000))
        excerpt = cleaned[:clean_limit]
        cp = (cfg.user_clean_prompt or _default_cleaning_prompt()).strip()
        try:
            llm_clean_summary = _deepseek_submit(cp + excerpt).strip()
        except Exception as e:
            llm_clean_summary = f"（Cloud 清洗失败：{e}）"

        # ══════════════════════════════════════════════════════════
        # 阶段 2：两条线并行 —— A) 规则+Skills  B) LLM上下文
        # ══════════════════════════════════════════════════════════
        q.put(("progress", 35, "规则扫描 + LLM 上下文并行处理…"))

        # 两路线程共享的结果容器
        line_a: dict = {}
        line_b: dict = {}

        def _line_a_rules_and_skills() -> None:
            """线 A：规则匹配 → Skills 智能判定"""
            rules = build_compiled_rules(
                env_json_path=None,
                user_json_path=_USER_RULES_PATH,
            )
            all_lines = cleaned.splitlines()
            buf: list[str] = []
            hit_lines = 0
            hit_categories: set[str] = set()
            hit_line_map: dict[str, list[int]] = {}
            hit_lines_with_text: list[tuple[int, str]] = []

            for i, line in enumerate(all_lines):
                hits = match_log_alerts_for_rules(line, rules)
                if not hits:
                    continue
                hit_lines += 1
                hit_lines_with_text.append((i, line))
                for h in hits:
                    hit_categories.add(h["cat"])
                    hit_line_map.setdefault(h["cat"], []).append(i)
                if len(buf) < 120:
                    labs = "、".join(h["label"] for h in hits)
                    buf.append(f"行 {i+1}: [{labs}] {line[:400]}{'…' if len(line) > 400 else ''}")

            if not buf:
                rule_block = "（本文件未命中内置/自定义规则关键字模式）"
            else:
                rule_block = "\n".join(buf)

            skill_block = ""
            if hit_categories:
                q.put(("progress", 38, "Skills 智能判定…"))
                try:
                    from utils.cursor_skills import run_all_skills, format_skill_results
                    skill_results = run_all_skills(
                        all_lines=all_lines,
                        hit_categories=hit_categories,
                        hit_line_map=hit_line_map,
                        hit_lines_with_text=hit_lines_with_text,
                        llm_fn=_deepseek_submit,
                        max_skills=5,
                        progress_fn=lambda msg: q.put(("progress", 42, msg)),
                    )
                    skill_block = format_skill_results(skill_results)
                except Exception as e:
                    skill_block = f"（Skills 判定跳过：{e}）"

            line_a["rule_block"] = rule_block
            line_a["skill_block"] = skill_block
            line_a["hit_lines"] = hit_lines

        def _line_b_llm_context() -> None:
            """线 B：解析异常行 → 结构化材料 + 历史判别经验"""
            incidents: list[Incident] = parse_report_text(cleaned, cfg.file_path.name)

            budget = max(4000, min(cfg.max_chars, 240_000))
            if incidents:
                material, truncated = build_material_for_prompt(incidents, budget)
            else:
                material, truncated = _raw_material(cleaned, budget, cfg.file_path.name)

            learning_ctx = ""
            try:
                from utils.db_manager import get_learning_context
                learning_ctx = get_learning_context()
            except Exception:
                pass

            line_b["incidents"] = incidents
            line_b["material"] = material
            line_b["truncated"] = truncated
            line_b["structured"] = bool(incidents)
            line_b["learning_ctx"] = learning_ctx

        t_a = threading.Thread(target=_line_a_rules_and_skills, daemon=True)
        t_b = threading.Thread(target=_line_b_llm_context, daemon=True)
        t_a.start()
        t_b.start()
        t_a.join()
        t_b.join()

        # ══════════════════════════════════════════════════════════
        # 阶段 3：合并两线产物 → 组装最终材料
        # ══════════════════════════════════════════════════════════
        q.put(("progress", 65, "组装分析材料…"))

        prefix_parts: list[str] = []

        # 清洗输出（最前置，作为全局上下文）
        prefix_parts.append(
            "## Cloud 清洗输出（前置）\n\n" + (llm_clean_summary or "（Cloud 未返回内容）")
        )

        # 用户侧重点
        if cfg.user_analysis_notes.strip():
            prefix_parts.append(
                "## 用户侧重点（清洗/分析）\n\n" + cfg.user_analysis_notes.strip()
            )

        # 规则命中摘要
        prefix_parts.append(
            f"## 规则命中摘要（共约 {line_a.get('hit_lines', 0)} 行触发规则）\n\n"
            + line_a.get("rule_block", "")
        )

        # Skills 判定结果
        skill_block = line_a.get("skill_block", "")
        if skill_block:
            prefix_parts.append(skill_block)

        # 历史判别经验（持续学习）
        learning_ctx = line_b.get("learning_ctx", "")
        if learning_ctx:
            prefix_parts.append(learning_ctx)

        material = "\n\n".join(prefix_parts) + "\n\n---\n\n" + line_b["material"]

        # ══════════════════════════════════════════════════════════
        # 阶段 4：LLM 总结 + 入库
        # ══════════════════════════════════════════════════════════
        structured = line_b.get("structured", False)
        truncated = line_b.get("truncated", False)
        prompt = build_summarize_prompt_serial(
            material,
            [cfg.file_path.name],
            truncated,
            structured,
        )

        q.put(("progress", 75, "调用 DeepSeek 生成总结（可能较久）…"))
        summary = _deepseek_submit(prompt)
        if not summary or not summary.strip():
            q.put(("err", "DeepSeek 返回空内容。"))
            return

        q.put(("progress", 95, "保存 Bug 到数据库…"))
        db_count = 0
        try:
            from utils.db_manager import save_bugs_from_summary
            db_count = save_bugs_from_summary(summary, cfg.file_path.name)
        except Exception as db_err:
            q.put(("progress", 95, f"数据库写入跳过（{db_err}）"))

        q.put(("progress", 100, "完成"))
        incidents = line_b.get("incidents", [])
        meta = (
            f"=== meta ===\n"
            f"generated: {datetime.now().isoformat(timespec='seconds')}\n"
            f"source: {cfg.file_path}\n"
            f"parsed_incidents: {len(incidents)}\n"
            f"structured_mode: {structured}\n"
            f"material_truncated: {truncated}\n"
            f"rule_hit_lines: {line_a.get('hit_lines', 0)}\n"
            f"bugs_saved_to_db: {db_count}\n\n"
            f"=== LLM summary ===\n\n"
        )
        q.put(("ok", meta + summary.strip()))
    except Exception as e:
        q.put(("err", f"{e}\n\n{traceback.format_exc()}"))


# ═══════════════════════════════════════════════════════════════
# PyQt6 UI 层
# ═══════════════════════════════════════════════════════════════

from PyQt6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QDialogButtonBox,
    QFileDialog,
    QFormLayout,
    QGridLayout,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMenu,
    QMenuBar,
    QMessageBox,
    QProgressBar,
    QPushButton,
    QSplitter,
    QStatusBar,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)
from PyQt6.QtCore import Qt, QThread, pyqtSignal, QObject
from PyQt6.QtGui import QAction, QFont, QFontDatabase


# ── QSS 全局样式 ──────────────────────────────────────────────

QSS = """
/* ── 全局 ───────────────────────────────────────────── */
QMainWindow {
    background-color: #f0f2f5;
}
QDialog {
    background-color: #f0f2f5;
}

/* ── 卡片 / 分组框 ──────────────────────────────────── */
QGroupBox {
    background-color: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    margin-top: 14px;
    padding: 20px 16px 14px 16px;
    font-weight: 600;
    font-size: 13px;
    color: #374151;
}
QGroupBox::title {
    subcontrol-origin: margin;
    left: 14px;
    padding: 0 8px;
    color: #6b7280;
}

/* ── 标签 ───────────────────────────────────────────── */
QLabel#pathLabel {
    color: #374151;
    font-size: 13px;
    padding: 0;
}
QLabel#mutedLabel {
    color: #9ca3af;
    font-size: 12px;
}

/* ── 按钮通用 ────────────────────────────────────────── */
QPushButton {
    border: 1px solid #d1d5db;
    border-radius: 7px;
    padding: 7px 16px;
    background-color: #ffffff;
    color: #374151;
    font-size: 13px;
    font-weight: 500;
}
QPushButton:hover {
    background-color: #f9fafb;
    border-color: #9ca3af;
}
QPushButton:pressed {
    background-color: #f3f4f6;
}

/* ── 主操作按钮 ─────────────────────────────────────── */
QPushButton#accentBtn {
    background-color: #4f46e5;
    color: #ffffff;
    border: none;
    font-weight: 600;
    font-size: 13px;
    padding: 8px 20px;
    border-radius: 7px;
}
QPushButton#accentBtn:hover {
    background-color: #4338ca;
}
QPushButton#accentBtn:pressed {
    background-color: #3730a3;
}
QPushButton#accentBtn:disabled {
    background-color: #a5b4fc;
    color: #e0e7ff;
}

/* ── 文本编辑区 ─────────────────────────────────────── */
QTextEdit {
    background-color: #fafbfc;
    color: #1f2937;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    padding: 10px 12px;
    font-size: 13px;
    selection-background-color: #c7d2fe;
    selection-color: #1e1b4b;
}
QTextEdit:focus {
    border-color: #a5b4fc;
}

/* ── 输入框 ─────────────────────────────────────────── */
QLineEdit {
    border: 1px solid #d1d5db;
    border-radius: 7px;
    padding: 7px 10px;
    background-color: #ffffff;
    color: #1f2937;
    font-size: 13px;
}
QLineEdit:focus {
    border-color: #4f46e5;
    background-color: #fafafe;
}
QLineEdit[echoMode="2"] {
    font-family: "Consolas", monospace;
}

/* ── 下拉框 ─────────────────────────────────────────── */
QComboBox {
    border: 1px solid #d1d5db;
    border-radius: 7px;
    padding: 6px 10px;
    background-color: #ffffff;
    color: #1f2937;
    font-size: 13px;
    min-width: 80px;
}
QComboBox:hover {
    border-color: #9ca3af;
}
QComboBox:focus {
    border-color: #4f46e5;
}
QComboBox::drop-down {
    subcontrol-origin: padding;
    subcontrol-position: top right;
    width: 24px;
    border: none;
    border-left: 1px solid #e5e7eb;
}
QComboBox QAbstractItemView {
    background-color: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 6px;
    padding: 4px;
    selection-background-color: #eef2ff;
    selection-color: #1f2937;
}

/* ── 树形列表 ────────────────────────────────────────── */
QTreeWidget {
    background-color: #ffffff;
    color: #374151;
    border: 1px solid #e5e7eb;
    border-radius: 8px;
    alternate-background-color: #fafbfc;
    selection-background-color: #eef2ff;
    selection-color: #312e81;
    outline: none;
    font-size: 13px;
}
QTreeWidget::item {
    padding: 5px 6px;
    border-bottom: 1px solid #f3f4f6;
}
QTreeWidget::item:hover {
    background-color: #f5f6ff;
}
QHeaderView::section {
    background-color: #f9fafb;
    color: #6b7280;
    border: none;
    border-bottom: 1px solid #e5e7eb;
    border-right: 1px solid #f3f4f6;
    padding: 7px 8px;
    font-weight: 600;
    font-size: 12px;
    letter-spacing: 0.01em;
}

/* ── 进度条 ─────────────────────────────────────────── */
QProgressBar {
    border: none;
    border-radius: 4px;
    background-color: #e5e7eb;
    height: 6px;
    text-align: center;
    font-size: 0px;
}
QProgressBar::chunk {
    background-color: #4f46e5;
    border-radius: 4px;
}

/* ── 菜单栏 ─────────────────────────────────────────── */
QMenuBar {
    background-color: #ffffff;
    border-bottom: 1px solid #f3f4f6;
    padding: 3px 6px;
    font-size: 13px;
}
QMenuBar::item {
    padding: 6px 12px;
    border-radius: 6px;
    color: #4b5563;
}
QMenuBar::item:selected {
    background-color: #f3f4f6;
    color: #1f2937;
}
QMenu {
    background-color: #ffffff;
    border: 1px solid #e5e7eb;
    border-radius: 10px;
    padding: 6px;
}
QMenu::item {
    padding: 7px 32px 7px 14px;
    border-radius: 6px;
    font-size: 13px;
    color: #374151;
}
QMenu::item:selected {
    background-color: #f3f4f6;
    color: #1f2937;
}
QMenu::separator {
    height: 1px;
    background: #f3f4f6;
    margin: 4px 8px;
}

/* ── 状态栏 ─────────────────────────────────────────── */
QStatusBar {
    background-color: #f0f2f5;
    color: #9ca3af;
    font-size: 12px;
    border-top: 1px solid #e5e7eb;
    padding: 2px 10px;
}

/* ── 分割器 ─────────────────────────────────────────── */
QSplitter::handle {
    background-color: transparent;
}
QSplitter::handle:vertical {
    height: 1px;
}

/* ── 滚动条 ─────────────────────────────────────────── */
QScrollBar:vertical {
    background: transparent;
    width: 8px;
    margin: 0;
}
QScrollBar::handle:vertical {
    background: #d1d5db;
    border-radius: 4px;
    min-height: 30px;
}
QScrollBar::handle:vertical:hover {
    background: #9ca3af;
}
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
    height: 0;
}
QScrollBar:horizontal {
    background: transparent;
    height: 8px;
}
QScrollBar::handle:horizontal {
    background: #d1d5db;
    border-radius: 4px;
    min-width: 30px;
}
QScrollBar::handle:horizontal:hover {
    background: #9ca3af;
}
QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
    width: 0;
}

/* ── 提示文本 ──────────────────────────────────────── */
QTextEdit[placeholderText], QLineEdit[placeholderText] {
    color: #d1d5db;
}
"""


# ── 信号适配器：让 _run_analyze 的 queue.put() 转为 Qt 信号 ──

class _QueueSignalAdapter(QObject):
    progress = pyqtSignal(int, str)
    result = pyqtSignal(str)
    error = pyqtSignal(str)

    def put(self, item: tuple) -> None:
        kind, *rest = item
        if kind == "progress":
            pct, msg = rest
            self.progress.emit(pct, msg)
        elif kind == "ok":
            self.result.emit(rest[0])
        elif kind == "err":
            self.error.emit(rest[0])


# ── 工作线程 ──────────────────────────────────────────────────

class AnalysisWorker(QThread):
    def __init__(self, cfg: WorkerConfig, parent: QObject | None = None):
        super().__init__(parent)
        self.cfg = cfg
        self._adapter = _QueueSignalAdapter()

    @property
    def progress_signal(self):
        return self._adapter.progress

    @property
    def result_signal(self):
        return self._adapter.result

    @property
    def error_signal(self):
        return self._adapter.error

    def run(self) -> None:
        _run_analyze(self.cfg, self._adapter)


class ImportWorker(QThread):
    finished = pyqtSignal(list, str)

    def __init__(self, src_path: Path, parent: QObject | None = None):
        super().__init__(parent)
        self._src_path = src_path

    def run(self) -> None:
        plain = _tabular_rules_source_to_plain(self._src_path)
        llm_text = _deepseek_submit(_rule_import_prompt(plain))
        new_items = _rules_from_llm_response(llm_text)
        if not new_items:
            raise ValueError("模型未返回任何可用规则。")
        self.finished.emit(new_items, self._src_path.name)


# ── 对话框：DeepSeek API 配置 ─────────────────────────────────

class DeepSeekConfigDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("DeepSeek API 配置")
        self.setMinimumWidth(500)
        self.setMinimumHeight(200)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(14)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self._api_key_edit = QLineEdit()
        self._api_key_edit.setEchoMode(QLineEdit.EchoMode.Password)
        self._api_key_edit.setPlaceholderText("sk-...")
        self._api_key_edit.setText(os.environ.get("DEEPSEEK_API_KEY", ""))
        self._api_key_edit.setMinimumHeight(34)

        self._model_edit = QLineEdit()
        self._model_edit.setPlaceholderText("deepseek-chat")
        self._model_edit.setText(os.environ.get("DEEPSEEK_MODEL", "deepseek-chat"))
        self._model_edit.setMinimumHeight(34)

        form.addRow("API Key ：", self._api_key_edit)
        form.addRow("模型：", self._model_edit)
        layout.addLayout(form)

        hint = QLabel("保存后写入项目 .env 文件；已运行的任务需下次「开始分析」生效。")
        hint.setWordWrap(True)
        hint.setObjectName("mutedLabel")
        layout.addWidget(hint)

        layout.addStretch()

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        b = buttons.button(QDialogButtonBox.StandardButton.Save)
        if b:
            b.setObjectName("accentBtn")
            b.setMinimumHeight(34)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _save(self) -> None:
        apply_deepseek_env_to_dotenv({
            "DEEPSEEK_API_KEY": self._api_key_edit.text().strip(),
            "DEEPSEEK_MODEL": self._model_edit.text().strip() or "deepseek-chat",
        })
        QMessageBox.information(self, "配置", "已保存到 .env。")
        self.accept()


# ── 对话框：清洗与分析材料 ─────────────────────────────────────

class CleanSettingsDialog(QDialog):
    def __init__(
        self,
        analysis_notes: str,
        clean_prompt: str,
        parent: QWidget | None = None,
    ):
        super().__init__(parent)
        self.setWindowTitle("清洗与分析材料")
        self.setMinimumSize(540, 540)
        self.resize(640, 580)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # 分析侧重点
        notes_label = QLabel("分析侧重点（可选）")
        notes_label.setStyleSheet("font-weight: 600; font-size: 13px; color: #374151;")
        layout.addWidget(notes_label)

        self._notes_edit = QTextEdit()
        self._notes_edit.setAcceptRichText(False)
        self._notes_edit.setPlaceholderText("输入分析侧重点，将作为上下文注入 LLM 提示…")
        self._notes_edit.setPlainText(analysis_notes)
        self._notes_edit.setMaximumHeight(90)
        layout.addWidget(self._notes_edit)

        # Cloud 清洗 prompt
        clean_label = QLabel("Cloud 清洗 prompt 前缀")
        clean_label.setStyleSheet("font-weight: 600; font-size: 13px; color: #374151;")
        layout.addWidget(clean_label)

        self._clean_edit = QTextEdit()
        self._clean_edit.setAcceptRichText(False)
        mono_font = QFont(QFontDatabase.systemFont(QFontDatabase.SystemFont.FixedFont))
        mono_font.setPointSize(10)
        self._clean_edit.setFont(mono_font)
        self._clean_edit.setPlainText(clean_prompt)
        layout.addWidget(self._clean_edit, stretch=1)

        # 体量设置
        env_group = QGroupBox("材料与节选体量（写入项目 .env）")
        env_form = QFormLayout(env_group)
        env_form.setSpacing(8)
        env_form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self._summary_chars = QLineEdit()
        self._summary_chars.setText(os.environ.get("SUMMARY_MAX_CHARS", "90000"))
        self._summary_chars.setMaximumWidth(120)
        self._summary_chars.setMinimumHeight(32)

        self._clean_chars = QLineEdit()
        self._clean_chars.setText(os.environ.get("CLEAN_MAX_CHARS", "24000"))
        self._clean_chars.setMaximumWidth(120)
        self._clean_chars.setMinimumHeight(32)

        env_form.addRow("SUMMARY_MAX_CHARS：", self._summary_chars)
        env_form.addRow("CLEAN_MAX_CHARS：", self._clean_chars)
        hint = QLabel("留空则删除 .env 中该键，下次使用内置默认。")
        hint.setObjectName("mutedLabel")
        env_form.addRow("", hint)

        layout.addWidget(env_group)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Close
        )
        b = buttons.button(QDialogButtonBox.StandardButton.Save)
        if b:
            b.setObjectName("accentBtn")
            b.setMinimumHeight(34)
        c = buttons.button(QDialogButtonBox.StandardButton.Close)
        if c:
            c.setMinimumHeight(34)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

        self._analysis_notes = analysis_notes
        self._clean_prompt_body = clean_prompt

    def _save(self) -> None:
        s_raw = self._summary_chars.text().strip()
        c_raw = self._clean_chars.text().strip()

        if s_raw:
            try:
                v = int(s_raw)
                if v < 4000 or v > 240_000:
                    QMessageBox.warning(self, "校验", "SUMMARY_MAX_CHARS 建议在 4000～240000。")
                    return
            except ValueError:
                QMessageBox.warning(self, "校验", "SUMMARY_MAX_CHARS 须为整数或留空。")
                return
        if c_raw:
            try:
                cv = int(c_raw)
                if cv < 1000 or cv > 80_000:
                    QMessageBox.warning(self, "校验", "CLEAN_MAX_CHARS 建议在 1000～80000。")
                    return
            except ValueError:
                QMessageBox.warning(self, "校验", "CLEAN_MAX_CHARS 须为整数或留空。")
                return

        self._analysis_notes = self._notes_edit.toPlainText().rstrip("\n")
        self._clean_prompt_body = self._clean_edit.toPlainText().rstrip("\n")
        apply_deepseek_env_to_dotenv({
            "SUMMARY_MAX_CHARS": s_raw,
            "CLEAN_MAX_CHARS": c_raw,
        })
        QMessageBox.information(self, "清洗与分析材料", "已保存文本与 .env 项。")

    def get_notes(self) -> str:
        return self._analysis_notes

    def get_clean_prompt(self) -> str:
        return self._clean_prompt_body

    def refresh_fields(self, analysis_notes: str, clean_prompt: str) -> None:
        self._notes_edit.setPlainText(analysis_notes)
        self._clean_edit.setPlainText(clean_prompt)
        self._summary_chars.setText(os.environ.get("SUMMARY_MAX_CHARS", "90000"))
        self._clean_chars.setText(os.environ.get("CLEAN_MAX_CHARS", "24000"))


# ── 对话框：添加自定义规则 ─────────────────────────────────────

class AddRuleDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("添加自定义规则")
        self.setMinimumWidth(460)
        self.setWindowModality(Qt.WindowModality.ApplicationModal)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        form = QFormLayout()
        form.setSpacing(10)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self._pri_edit = QLineEdit("5")
        self._pri_edit.setMaximumWidth(80)
        self._pri_edit.setMinimumHeight(32)

        self._cat_edit = QLineEdit()
        self._cat_edit.setMinimumHeight(32)
        self._lbl_edit = QLineEdit()
        self._lbl_edit.setMinimumHeight(32)
        self._pat_edit = QLineEdit()
        self._pat_edit.setMinimumHeight(32)

        form.addRow("优先级：", self._pri_edit)
        form.addRow("category：", self._cat_edit)
        form.addRow("显示名：", self._lbl_edit)
        form.addRow("pattern：", self._pat_edit)

        hint = QLabel("优先级数字越小越优先；category 为唯一键，不可与已有规则重复。")
        hint.setWordWrap(True)
        hint.setObjectName("mutedLabel")
        form.addRow("", hint)

        layout.addLayout(form)
        layout.addStretch()

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel
        )
        b = buttons.button(QDialogButtonBox.StandardButton.Ok)
        if b:
            b.setObjectName("accentBtn")
            b.setMinimumHeight(34)
        buttons.accepted.connect(self._validate_and_accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _validate_and_accept(self) -> None:
        try:
            p = int(self._pri_edit.text().strip() or "5")
        except ValueError:
            QMessageBox.warning(self, "校验", "优先级必须是整数。")
            return
        cat = self._cat_edit.text().strip()
        pat = self._pat_edit.text().strip()
        lbl = self._lbl_edit.text().strip() or cat
        if not cat or not pat:
            QMessageBox.warning(self, "校验", "请填写 category 与 pattern。")
            return
        try:
            re.compile(pat, re.I)
        except re.error as e:
            QMessageBox.warning(self, "校验", f"正则无效：{e}")
            return
        items = load_user_rules_raw(_USER_RULES_PATH)
        items.append({"priority": p, "category": cat, "label": lbl, "pattern": pat})
        save_user_rules_raw(_USER_RULES_PATH, items)
        self.accept()


# ── 对话框：串口匹配规则管理 ──────────────────────────────────

class RulesDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("串口匹配规则")
        self.setMinimumSize(680, 420)
        self.resize(780, 480)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # 工具栏
        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        add_btn = QPushButton("+ 添加规则")
        add_btn.setMinimumHeight(32)
        add_btn.clicked.connect(self._add_rule)
        del_btn = QPushButton("删除选中")
        del_btn.setMinimumHeight(32)
        del_btn.clicked.connect(self._del_rule)
        refresh_btn = QPushButton("刷新")
        refresh_btn.setMinimumHeight(32)
        refresh_btn.clicked.connect(self.refresh_table)

        toolbar.addWidget(add_btn)
        toolbar.addWidget(del_btn)
        toolbar.addWidget(refresh_btn)
        toolbar.addStretch()
        layout.addLayout(toolbar)

        # 规则表格
        self._tree = QTreeWidget()
        self._tree.setColumnCount(5)
        self._tree.setHeaderLabels(["来源", "优先级", "category", "显示名", "pattern"])
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(False)
        self._tree.setSelectionBehavior(QTreeWidget.SelectionBehavior.SelectRows)
        self._tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)

        header = self._tree.header()
        header.setStretchLastSection(True)
        header.resizeSection(0, 60)
        header.resizeSection(1, 56)
        header.resizeSection(2, 120)
        header.resizeSection(3, 160)

        layout.addWidget(self._tree, stretch=1)

        # 关闭按钮
        btn_row = QHBoxLayout()
        btn_row.addStretch()
        close_btn = QPushButton("确定")
        close_btn.setMinimumHeight(32)
        close_btn.setMinimumWidth(80)
        close_btn.clicked.connect(self.hide)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self.refresh_table()

    def refresh_table(self) -> None:
        self._tree.clear()
        for pri, cat, lbl, pat in RAW_ALERT_RULE_DEFINITIONS:
            item = QTreeWidgetItem(["内置", str(pri), cat, lbl, pat])
            self._tree.addTopLevelItem(item)
        user_items = load_user_rules_raw(_USER_RULES_PATH)
        for it in user_items:
            item = QTreeWidgetItem([
                "自定义", str(it["priority"]), it["category"], it["label"], it["pattern"],
            ])
            self._tree.addTopLevelItem(item)

    def _add_rule(self) -> None:
        dlg = AddRuleDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh_table()

    def _del_rule(self) -> None:
        sel = self._tree.selectedItems()
        if not sel:
            return
        item = sel[0]
        if item.text(0) != "自定义":
            QMessageBox.information(self, "提示", "只能删除「自定义」规则。")
            return
        idx = self._tree.indexOfTopLevelItem(item)
        items = load_user_rules_raw(_USER_RULES_PATH)
        if 0 <= idx < len(items):
            items.pop(idx)
            save_user_rules_raw(_USER_RULES_PATH, items)
            self.refresh_table()


# ── 对话框：Bug 记录管理 ──────────────────────────────────────

class BugRecordsDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Bug 记录（PostgreSQL）")
        self.setMinimumSize(900, 500)
        self.resize(1000, 580)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # 顶部筛选栏
        top_bar = QHBoxLayout()
        top_bar.setSpacing(10)

        top_bar.addWidget(QLabel("筛选判别："))
        self._filter_combo = QComboBox()
        self._filter_combo.addItems(["全部", "待定", "确认", "误报", "忽略"])
        self._filter_combo.setMinimumWidth(90)
        self._filter_combo.currentTextChanged.connect(self.refresh_table)
        top_bar.addWidget(self._filter_combo)

        refresh_btn = QPushButton("刷新")
        refresh_btn.setMinimumHeight(32)
        refresh_btn.clicked.connect(self.refresh_table)
        top_bar.addWidget(refresh_btn)

        top_bar.addStretch()

        self._count_label = QLabel()
        self._count_label.setObjectName("mutedLabel")
        top_bar.addWidget(self._count_label)
        layout.addLayout(top_bar)

        # Bug 表格
        self._tree = QTreeWidget()
        self._tree.setColumnCount(6)
        self._tree.setHeaderLabels(["ID", "Bug编号", "标题", "严重级别", "来源文件", "人工判别"])
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(False)
        self._tree.setSelectionBehavior(QTreeWidget.SelectionBehavior.SelectRows)
        self._tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        self._tree.itemSelectionChanged.connect(self._on_select)

        header = self._tree.header()
        header.setStretchLastSection(False)
        header.resizeSection(0, 50)
        header.resizeSection(1, 84)
        header.resizeSection(2, 300)
        header.resizeSection(3, 72)
        header.resizeSection(4, 240)
        header.resizeSection(5, 80)

        layout.addWidget(self._tree, stretch=1)

        # 审核面板
        review_card = QWidget()
        review_card.setStyleSheet("""
            QWidget#reviewCard {
                background: #fafbfc;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }
        """)
        review_card.setObjectName("reviewCard")
        review_layout = QVBoxLayout(review_card)
        review_layout.setContentsMargins(14, 12, 14, 12)
        review_layout.setSpacing(8)

        row1 = QHBoxLayout()
        row1.setSpacing(10)

        row1.addWidget(QLabel("判别："))
        self._verdict_combo = QComboBox()
        self._verdict_combo.addItems(["待定", "确认", "误报", "忽略"])
        self._verdict_combo.setMinimumWidth(80)
        row1.addWidget(self._verdict_combo)

        row1.addWidget(QLabel("备注："))
        self._notes_edit = QLineEdit()
        self._notes_edit.setPlaceholderText("输入审核备注…")
        self._notes_edit.setMinimumHeight(32)
        row1.addWidget(self._notes_edit, stretch=1)

        save_btn = QPushButton("保存判别")
        save_btn.setObjectName("accentBtn")
        save_btn.setMinimumHeight(32)
        save_btn.clicked.connect(self._save_verdict)
        row1.addWidget(save_btn)

        del_btn = QPushButton("删除记录")
        del_btn.setMinimumHeight(32)
        del_btn.clicked.connect(self._delete_record)
        row1.addWidget(del_btn)

        review_layout.addLayout(row1)

        self._detail_label = QLabel("选中一行查看详情")
        self._detail_label.setWordWrap(True)
        self._detail_label.setObjectName("mutedLabel")
        review_layout.addWidget(self._detail_label)

        layout.addWidget(review_card)

        self._selected_bug_id: int | None = None

        try:
            from utils.db_manager import init_db
            init_db()
        except Exception:
            pass
        self.refresh_table()

    def refresh_table(self) -> None:
        self._tree.clear()
        try:
            from utils.db_manager import list_bugs, count_bugs
            flt = self._filter_combo.currentText()
            rows = list_bugs(limit=500, verdict=flt if flt != "全部" else None)
            for r in rows:
                src = r.get("source_file", "")
                if len(src) > 40:
                    src = "…" + src[-38:]
                item = QTreeWidgetItem([
                    str(r["id"]),
                    r.get("bug_no", ""),
                    (r.get("title", "") or "")[:60],
                    r.get("severity", ""),
                    src,
                    r.get("human_verdict", "待定"),
                ])
                self._tree.addTopLevelItem(item)
            stats = count_bugs()
            self._count_label.setText(
                f"共 {stats.get('total',0)} 条 | "
                f"待定 {stats.get('待定',0)} · 确认 {stats.get('确认',0)} · "
                f"误报 {stats.get('误报',0)} · 忽略 {stats.get('忽略',0)}"
            )
        except Exception as e:
            self._count_label.setText(f"加载失败：{e}")

    def _on_select(self) -> None:
        sel = self._tree.selectedItems()
        if not sel:
            self._selected_bug_id = None
            return
        bug_id = int(sel[0].text(0))
        self._selected_bug_id = bug_id
        try:
            from utils.db_manager import get_bug
            b = get_bug(bug_id)
            if not b:
                return
            idx = self._verdict_combo.findText(b.get("human_verdict", "待定"))
            if idx >= 0:
                self._verdict_combo.setCurrentIndex(idx)
            self._notes_edit.setText(b.get("human_notes", ""))
            self._detail_label.setText(
                f"[{b.get('bug_no','')}] {b.get('title','')}\n"
                f"类型：{b.get('bug_type','')}  时间：{b.get('log_time','')}\n"
                f"结论：{(b.get('conclusion','') or '')[:200]}"
            )
        except Exception:
            pass

    def _save_verdict(self) -> None:
        if self._selected_bug_id is None:
            QMessageBox.information(self, "提示", "请先选中一条 Bug。")
            return
        try:
            from utils.db_manager import update_verdict
            ok = update_verdict(
                self._selected_bug_id,
                self._verdict_combo.currentText(),
                self._notes_edit.text(),
            )
            if ok:
                self.refresh_table()
                # re-select
                for i in range(self._tree.topLevelItemCount()):
                    item = self._tree.topLevelItem(i)
                    if item and item.text(0) == str(self._selected_bug_id):
                        self._tree.setCurrentItem(item)
                        break
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e)[:500])

    def _delete_record(self) -> None:
        if self._selected_bug_id is None:
            return
        reply = QMessageBox.question(
            self, "确认", f"删除 Bug #{self._selected_bug_id}？",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
        )
        if reply != QMessageBox.StandardButton.Yes:
            return
        try:
            from utils.db_manager import delete_bug
            delete_bug(self._selected_bug_id)
            self._selected_bug_id = None
            self.refresh_table()
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e)[:500])


# ── 对话框：添加自定义 Skill ──────────────────────────────────

class AddSkillDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("添加自定义 Skill")
        self.setMinimumSize(540, 520)
        self.resize(560, 560)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(10)

        form = QFormLayout()
        form.setSpacing(8)
        form.setLabelAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)

        self._id_edit = QLineEdit()
        self._id_edit.setMinimumHeight(32)
        self._name_edit = QLineEdit()
        self._name_edit.setMinimumHeight(32)
        self._trigger_cats = QLineEdit()
        self._trigger_cats.setPlaceholderText("用逗号分隔，如：crash,memory")
        self._trigger_cats.setMinimumHeight(32)
        self._trigger_kws = QLineEdit()
        self._trigger_kws.setPlaceholderText("用逗号分隔")
        self._trigger_kws.setMinimumHeight(32)
        self._ctx_lines = QLineEdit("25")
        self._ctx_lines.setMinimumHeight(32)
        self._priority = QLineEdit("5")
        self._priority.setMinimumHeight(32)

        form.addRow("Skill ID：", self._id_edit)
        form.addRow("名称：", self._name_edit)
        form.addRow("触发类别：", self._trigger_cats)
        form.addRow("触发关键词：", self._trigger_kws)
        form.addRow("上下文行数：", self._ctx_lines)
        form.addRow("优先级：", self._priority)
        layout.addLayout(form)

        prompt_label = QLabel("Prompt 模板")
        prompt_label.setStyleSheet("font-weight: 600; font-size: 13px; color: #374151;")
        layout.addWidget(prompt_label)

        self._prompt_edit = QTextEdit()
        self._prompt_edit.setAcceptRichText(False)
        mono_font = QFont(QFontDatabase.systemFont(QFontDatabase.SystemFont.FixedFont))
        mono_font.setPointSize(10)
        self._prompt_edit.setFont(mono_font)
        self._prompt_edit.setPlainText(
            "你是XXX分析专家。以下日志片段包含可能的XXX异常。\n"
            "请严格按以下格式输出：\n"
            "【判定】确认异常 / 疑似误报 / 需更多上下文\n"
            "【置信度】高/中/低\n"
            "【关键发现】一句话概述\n"
            "【建议】一句话排查建议\n\n"
            "--- 日志片段 ---\n"
        )
        layout.addWidget(self._prompt_edit, stretch=1)

        buttons = QDialogButtonBox(
            QDialogButtonBox.StandardButton.Save | QDialogButtonBox.StandardButton.Cancel
        )
        b = buttons.button(QDialogButtonBox.StandardButton.Save)
        if b:
            b.setObjectName("accentBtn")
            b.setMinimumHeight(34)
        buttons.accepted.connect(self._save)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)

    def _save(self) -> None:
        sid = self._id_edit.text().strip()
        if not sid:
            QMessageBox.warning(self, "提示", "Skill ID 不能为空")
            return
        try:
            from utils.cursor_skills import add_user_skill
            add_user_skill({
                "id": sid,
                "name": self._name_edit.text().strip() or sid,
                "trigger_categories": [
                    c.strip()
                    for c in self._trigger_cats.text().split(",")
                    if c.strip()
                ],
                "trigger_keywords": [
                    k.strip()
                    for k in self._trigger_kws.text().split(",")
                    if k.strip()
                ],
                "context_lines": int(self._ctx_lines.text() or "25"),
                "priority": int(self._priority.text() or "5"),
                "prompt_template": self._prompt_edit.toPlainText().strip() + "\n",
                "enabled": True,
            })
            self.accept()
        except Exception as e:
            QMessageBox.critical(self, "保存失败", str(e)[:500])


# ── 对话框：Skills 管理 ───────────────────────────────────────

class SkillsManagerDialog(QDialog):
    def __init__(self, parent: QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Skills 管理")
        self.setMinimumSize(740, 420)
        self.resize(820, 500)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 16)
        layout.setSpacing(12)

        # 工具栏
        toolbar = QHBoxLayout()
        toolbar.setSpacing(8)

        refresh_btn = QPushButton("刷新")
        refresh_btn.setMinimumHeight(32)
        refresh_btn.clicked.connect(self.refresh_table)
        add_btn = QPushButton("+ 添加 Skill")
        add_btn.setMinimumHeight(32)
        add_btn.clicked.connect(self._add_skill)
        del_btn = QPushButton("删除选中")
        del_btn.setMinimumHeight(32)
        del_btn.clicked.connect(self._delete_skill)

        toolbar.addWidget(refresh_btn)
        toolbar.addWidget(add_btn)
        toolbar.addWidget(del_btn)
        toolbar.addStretch()

        self._count_label = QLabel()
        self._count_label.setObjectName("mutedLabel")
        toolbar.addWidget(self._count_label)
        layout.addLayout(toolbar)

        # Skills 表格
        self._tree = QTreeWidget()
        self._tree.setColumnCount(6)
        self._tree.setHeaderLabels(["ID", "名称", "触发类别", "优先级", "启用", "来源"])
        self._tree.setAlternatingRowColors(True)
        self._tree.setRootIsDecorated(False)
        self._tree.setSelectionBehavior(QTreeWidget.SelectionBehavior.SelectRows)
        self._tree.setSelectionMode(QTreeWidget.SelectionMode.SingleSelection)
        self._tree.itemDoubleClicked.connect(self._toggle_enabled)

        header = self._tree.header()
        header.setStretchLastSection(False)
        header.resizeSection(0, 140)
        header.resizeSection(1, 170)
        header.resizeSection(2, 210)
        header.resizeSection(3, 60)
        header.resizeSection(4, 54)
        header.resizeSection(5, 64)

        layout.addWidget(self._tree, stretch=1)

        hint = QLabel("提示：双击行切换启用/禁用；内置 Skill 不可删除，可双击禁用。")
        hint.setObjectName("mutedLabel")
        layout.addWidget(hint)

        self.refresh_table()

    def refresh_table(self) -> None:
        self._tree.clear()
        try:
            from utils.cursor_skills import get_all_skills
            skills = get_all_skills()
            for s in skills:
                cats = ", ".join(s.trigger_categories[:4])
                item = QTreeWidgetItem([
                    s.id, s.name, cats, str(s.priority),
                    "✓" if s.enabled else "✗",
                    "内置" if s.builtin else "自定义",
                ])
                self._tree.addTopLevelItem(item)
            self._count_label.setText(f"共 {len(skills)} 个 Skill")
        except Exception as e:
            self._count_label.setText(f"加载失败：{e}")

    def _toggle_enabled(self, item: QTreeWidgetItem) -> None:
        skill_id = item.text(0)
        try:
            from utils.cursor_skills import get_skill_by_id, toggle_skill
            s = get_skill_by_id(skill_id)
            if s:
                toggle_skill(skill_id, not s.enabled)
                self.refresh_table()
        except Exception:
            pass

    def _add_skill(self) -> None:
        dlg = AddSkillDialog(self)
        if dlg.exec() == QDialog.DialogCode.Accepted:
            self.refresh_table()

    def _delete_skill(self) -> None:
        sel = self._tree.selectedItems()
        if not sel:
            return
        skill_id = sel[0].text(0)
        try:
            from utils.cursor_skills import get_skill_by_id, delete_user_skill
            s = get_skill_by_id(skill_id)
            if s and s.builtin:
                QMessageBox.information(self, "提示", "内置 Skill 不可删除，可双击禁用。")
                return
            reply = QMessageBox.question(
                self, "确认", f"删除自定义 Skill「{skill_id}」？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return
            delete_user_skill(skill_id)
            self.refresh_table()
        except Exception as e:
            QMessageBox.critical(self, "删除失败", str(e)[:500])


# ── 主窗口 ────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("串口日志分析")
        self.setMinimumSize(520, 420)

        screen = QApplication.primaryScreen()
        if screen:
            sg = screen.availableGeometry()
            ww = min(780, max(520, int(sg.width() * 0.52)))
            wh = min(640, max(460, int(sg.height() * 0.62)))
            self.resize(ww, wh)
            self.move((sg.width() - ww) // 2, (sg.height() - wh) // 2)

        self.setStyleSheet(QSS)

        # 状态变量
        self._current_file: Path | None = None
        self._last_result = ""
        self._analysis_notes = ""
        self._clean_prompt_body = _default_cleaning_prompt()
        self._worker: AnalysisWorker | None = None
        self._import_worker: ImportWorker | None = None

        # 懒加载对话框
        self._deepseek_dlg: DeepSeekConfigDialog | None = None
        self._clean_dlg: CleanSettingsDialog | None = None
        self._rules_dlg: RulesDialog | None = None
        self._bug_dlg: BugRecordsDialog | None = None
        self._skills_dlg: SkillsManagerDialog | None = None

        self.setFont(QFont("Microsoft YaHei UI", 9))
        self._build_menu_bar()
        self._build_central_widget()
        self._build_status_bar()

    # ── 菜单栏 ─────────────────────────────────────────────────

    def _build_menu_bar(self) -> None:
        mb = self.menuBar()

        # 文件
        file_menu = mb.addMenu("文件")
        file_menu.addAction(QAction("打开…", self, triggered=self._pick_file))

        # 设置
        settings_menu = mb.addMenu("设置")
        settings_menu.addAction(QAction("DeepSeek API…", self, triggered=self._open_deepseek_config))
        settings_menu.addAction(QAction("清洗与分析材料…", self, triggered=self._open_clean_config))

        # 规则
        rules_menu = mb.addMenu("规则")
        rules_menu.addAction(QAction("串口匹配规则…", self, triggered=self._open_rules_config))
        rules_menu.addAction(QAction("从 Excel/CSV 导入规则…", self, triggered=self._import_rules_from_xlsx))

        # 数据库
        db_menu = mb.addMenu("数据库")
        db_menu.addAction(QAction("Bug 记录…", self, triggered=self._open_bug_records))
        db_menu.addAction(QAction("测试连接…", self, triggered=self._test_db_connection))

        # Skills
        skills_menu = mb.addMenu("Skills")
        skills_menu.addAction(QAction("管理 Skills…", self, triggered=self._open_skills_manager))

        # 直接命令
        mb.addAction(QAction("免责声明", self, triggered=self._show_disclaimer))
        mb.addAction(QAction("关于", self, triggered=self._show_about))

    # ── 中央区域 ───────────────────────────────────────────────

    def _build_central_widget(self) -> None:
        central = QWidget()
        root = QVBoxLayout(central)
        root.setContentsMargins(14, 10, 14, 10)
        root.setSpacing(10)

        # ── 顶部：文件路径条 ──────────────────────────────────
        path_bar = QWidget()
        path_bar.setStyleSheet("""
            QWidget#pathBar {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 8px;
            }
        """)
        path_bar.setObjectName("pathBar")
        path_row = QHBoxLayout(path_bar)
        path_row.setContentsMargins(14, 10, 14, 10)
        path_row.setSpacing(8)

        file_icon = QLabel("\U0001f4c4")
        file_icon.setStyleSheet("font-size: 16px; border: none; background: transparent;")
        path_row.addWidget(file_icon)

        self._path_label = QLabel("未选择文件（文件 → 打开…；规则见「规则」菜单，清洗见「设置」）")
        self._path_label.setObjectName("pathLabel")
        self._path_label.setWordWrap(True)
        self._path_label.setStyleSheet("border: none; background: transparent;")
        path_row.addWidget(self._path_label, stretch=1)

        root.addWidget(path_bar)

        # ── 主体：结果区 ──────────────────────────────────────
        result_card = QWidget()
        result_card.setStyleSheet("""
            QWidget#resultCard {
                background: #ffffff;
                border: 1px solid #e5e7eb;
                border-radius: 10px;
            }
        """)
        result_card.setObjectName("resultCard")
        result_layout = QVBoxLayout(result_card)
        result_layout.setContentsMargins(14, 14, 14, 12)
        result_layout.setSpacing(10)

        # 标题行
        header_row = QHBoxLayout()
        header_row.setSpacing(6)
        title_lbl = QLabel("分析结果")
        title_lbl.setStyleSheet(
            "font-weight: 600; font-size: 14px; color: #374151; border: none; background: transparent;"
        )
        header_row.addWidget(title_lbl)
        header_row.addStretch()

        self._char_count_lbl = QLabel("")
        self._char_count_lbl.setObjectName("mutedLabel")
        self._char_count_lbl.setStyleSheet("border: none; background: transparent;")
        header_row.addWidget(self._char_count_lbl)
        result_layout.addLayout(header_row)

        # 文本区
        self._result_edit = QTextEdit()
        self._result_edit.setReadOnly(True)
        self._result_edit.setPlaceholderText("分析结果将在此显示…")
        mono_font = QFont(QFontDatabase.systemFont(QFontDatabase.SystemFont.FixedFont))
        mono_font.setPointSize(10)
        self._result_edit.setFont(mono_font)
        result_layout.addWidget(self._result_edit, stretch=1)

        # 底部工具栏
        toolbar = QHBoxLayout()
        toolbar.setSpacing(10)

        self._run_btn = QPushButton("  开始分析")
        self._run_btn.setObjectName("accentBtn")
        self._run_btn.setMinimumHeight(36)
        self._run_btn.clicked.connect(self._run)
        toolbar.addWidget(self._run_btn)

        save_btn = QPushButton("保存 TXT…")
        save_btn.setMinimumHeight(36)
        save_btn.clicked.connect(self._save_result)
        toolbar.addWidget(save_btn)

        toolbar.addSpacing(10)

        self._progress_bar = QProgressBar()
        self._progress_bar.setMaximum(100)
        self._progress_bar.setValue(0)
        self._progress_bar.setTextVisible(False)
        self._progress_bar.setFixedHeight(6)
        self._progress_bar.setMinimumWidth(120)
        toolbar.addWidget(self._progress_bar, stretch=1)

        result_layout.addLayout(toolbar)
        root.addWidget(result_card, stretch=1)

        self.setCentralWidget(central)

    # ── 状态栏 ─────────────────────────────────────────────────

    def _build_status_bar(self) -> None:
        self._status_bar = QStatusBar()
        self.setStatusBar(self._status_bar)
        self._status_bar.showMessage("就绪  —  选择日志文件后点击「开始分析」")

    # ── 文件操作 ───────────────────────────────────────────────

    def _pick_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择串口日志",
            str(_ROOT),
            "日志文件 (*.log *.txt);;所有文件 (*.*)",
        )
        if path:
            self._current_file = Path(path)
            self._path_label.setText(str(self._current_file))

    # ── 菜单动作 ───────────────────────────────────────────────

    def _show_disclaimer(self) -> None:
        QMessageBox.information(self, "免责声明", DISCLAIMER_TEXT)

    def _show_about(self) -> None:
        QMessageBox.information(
            self, "关于",
            f"作者：{APP_AUTHOR}\n版本：{APP_VERSION}",
        )

    def _open_deepseek_config(self) -> None:
        if self._deepseek_dlg is None:
            self._deepseek_dlg = DeepSeekConfigDialog(self)
        self._deepseek_dlg.exec()

    def _open_clean_config(self) -> None:
        if self._clean_dlg is None:
            self._clean_dlg = CleanSettingsDialog(
                self._analysis_notes,
                self._clean_prompt_body,
                self,
            )
        else:
            self._clean_dlg.refresh_fields(self._analysis_notes, self._clean_prompt_body)
        if self._clean_dlg.exec() == QDialog.DialogCode.Accepted:
            self._analysis_notes = self._clean_dlg.get_notes()
            self._clean_prompt_body = self._clean_dlg.get_clean_prompt()

    def _open_rules_config(self) -> None:
        if self._rules_dlg is None:
            self._rules_dlg = RulesDialog(self)
        else:
            self._rules_dlg.refresh_table()
        self._rules_dlg.show()
        self._rules_dlg.raise_()
        self._rules_dlg.activateWindow()

    def _open_bug_records(self) -> None:
        if self._bug_dlg is None:
            self._bug_dlg = BugRecordsDialog(self)
        else:
            self._bug_dlg.refresh_table()
        self._bug_dlg.show()
        self._bug_dlg.raise_()
        self._bug_dlg.activateWindow()

    def _test_db_connection(self) -> None:
        try:
            from utils.db_manager import test_connection
            ok, msg = test_connection()
            if ok:
                QMessageBox.information(self, "数据库", msg)
            else:
                QMessageBox.critical(self, "数据库连接失败", msg)
        except Exception as e:
            QMessageBox.critical(self, "数据库", f"无法导入 db_manager：{e}")

    def _open_skills_manager(self) -> None:
        if self._skills_dlg is None:
            self._skills_dlg = SkillsManagerDialog(self)
        else:
            self._skills_dlg.refresh_table()
        self._skills_dlg.show()
        self._skills_dlg.raise_()
        self._skills_dlg.activateWindow()

    # ── 规则导入 ───────────────────────────────────────────────

    def _import_rules_from_xlsx(self) -> None:
        if self._import_worker is not None and self._import_worker.isRunning():
            QMessageBox.information(self, "提示", "已有规则导入任务在运行。")
            return
        if not os.environ.get("DEEPSEEK_API_KEY", "").strip():
            QMessageBox.warning(
                self, "未配置 DeepSeek",
                "请先在「设置 → DeepSeek API…」中填写 DEEPSEEK_API_KEY。",
            )
            return

        path, _ = QFileDialog.getOpenFileName(
            self,
            "选择压测用例等表格文件（Excel 或 CSV）",
            str(_ROOT),
            "Excel/CSV (*.xlsx *.xlsm *.csv);;Excel (*.xlsx *.xlsm);;CSV (*.csv);;所有文件 (*.*)",
        )
        if not path:
            return
        src_path = Path(path)
        if not src_path.is_file():
            QMessageBox.warning(self, "提示", "所选路径无效。")
            return
        if src_path.suffix.lower() not in (".csv", ".xlsx", ".xlsm"):
            QMessageBox.warning(self, "提示", "请选择 .xlsx、.xlsm 或 .csv 文件。")
            return

        self._status_bar.showMessage("正在读取表格并调用 DeepSeek 提取规则（可能较久）…")

        self._import_worker = ImportWorker(src_path)
        self._import_worker.finished.connect(self._on_import_finished)

        def on_error():
            QMessageBox.critical(self, "导入失败", "规则导入过程中出现错误。")

        self._import_worker.finished.connect(
            lambda items, name: self._finish_import(items, name)
        )
        self._import_worker.start()

    def _finish_import(self, new_items: list[dict], name: str) -> None:
        existing = load_user_rules_raw(_USER_RULES_PATH)
        reserved: set[str] = {x["category"] for x in existing} | set(_BUILTIN_RULE_CATEGORIES)
        added: list[dict] = []
        for it in new_items:
            c = it["category"]
            if c in reserved:
                base = c
                n = 2
                while f"{base}_{n}" in reserved:
                    n += 1
                c = f"{base}_{n}"
            it["category"] = c
            reserved.add(c)
            added.append(it)
        merged = existing + added
        save_user_rules_raw(_USER_RULES_PATH, merged)
        if self._rules_dlg is not None:
            self._rules_dlg.refresh_table()
        self._status_bar.showMessage(f"已导入 {len(added)} 条自定义规则。")
        QMessageBox.information(
            self, "导入完成",
            f"已从「{name}」经 DeepSeek 提取并追加 {len(added)} 条规则到 serial_rules_user.json。\n"
            f"（若 category 与已有或内置键冲突，已自动加后缀。）",
        )

    # ── 分析运行 ───────────────────────────────────────────────

    def _run(self) -> None:
        if self._worker is not None and self._worker.isRunning():
            QMessageBox.information(self, "提示", "已有任务在运行。")
            return
        if not self._current_file or not self._current_file.is_file():
            QMessageBox.warning(self, "提示", "请先选择有效的日志文件。")
            return

        mc = int(os.environ.get("SUMMARY_MAX_CHARS", "90000"))
        cfg = WorkerConfig(
            file_path=self._current_file,
            max_chars=mc,
            user_clean_prompt=self._clean_prompt_body,
            user_analysis_notes=self._analysis_notes,
        )

        self._result_edit.clear()
        self._progress_bar.setValue(0)
        self._run_btn.setEnabled(False)

        self._worker = AnalysisWorker(cfg)
        self._worker.progress_signal.connect(self._on_progress)
        self._worker.result_signal.connect(self._on_result)
        self._worker.error_signal.connect(self._on_error)
        self._worker.finished.connect(lambda: self._run_btn.setEnabled(True))
        self._worker.start()

    def _on_progress(self, pct: int, msg: str) -> None:
        self._progress_bar.setValue(pct)
        self._status_bar.showMessage(msg)
        if pct == 0:
            self._char_count_lbl.setText("")

    def _on_result(self, text: str) -> None:
        self._last_result = text
        self._result_edit.setPlainText(text)
        self._status_bar.showMessage("完成。")
        lines = text.count("\n") + 1
        chars = len(text)
        self._char_count_lbl.setText(f"{lines} 行 · {chars:,} 字符")

    def _on_error(self, err: str) -> None:
        self._result_edit.setPlainText(err)
        self._status_bar.showMessage("失败。")
        QMessageBox.critical(self, "错误", err[:800])

    # ── 保存结果 ───────────────────────────────────────────────

    def _save_result(self) -> None:
        if not self._last_result.strip():
            QMessageBox.information(self, "提示", "没有可保存的内容。")
            return
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        default = _ROOT / "bugs" / f"serial_gui_summary_{ts}.txt"
        path, _ = QFileDialog.getSaveFileName(
            self,
            "保存分析结果",
            str(default),
            "文本文件 (*.txt)",
        )
        if path:
            p = Path(path)
            p.parent.mkdir(parents=True, exist_ok=True)
            p.write_text(self._last_result, encoding="utf-8")
            self._status_bar.showMessage(f"已保存：{p}")


# ── 入口 ──────────────────────────────────────────────────────

def main() -> None:
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
